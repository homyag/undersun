from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods, require_POST
from django.core.paginator import Paginator
from django.db.models import Q
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
import json
import threading

from .models import ImportFile, ImportLog, PropertyImportMapping
from .forms import ImportFileForm, MappingForm, ImportPreviewForm, BulkActionForm
from .services import ImportProcessor, ExcelParser


def is_staff_user(user):
    """Проверка, что пользователь является сотрудником"""
    return user.is_staff


@login_required
@user_passes_test(is_staff_user)
def import_dashboard(request):
    """Главная страница импорта данных"""
    # Получаем статистику
    recent_imports = ImportFile.objects.select_related('uploaded_by').order_by('-created_at')[:5]
    
    statistics = {
        'total_imports': ImportFile.objects.count(),
        'processing_imports': ImportFile.objects.filter(status='processing').count(),
        'completed_imports': ImportFile.objects.filter(status='completed').count(),
        'failed_imports': ImportFile.objects.filter(status='failed').count(),
        'total_mappings': PropertyImportMapping.objects.count(),
    }
    
    context = {
        'recent_imports': recent_imports,
        'statistics': statistics,
    }
    
    return render(request, 'data_import/dashboard.html', context)


@login_required
@user_passes_test(is_staff_user)
def upload_file(request):
    """Загрузка файла для импорта"""
    if request.method == 'POST':
        form = ImportFileForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            import_file = form.save()
            messages.success(request, _('Файл успешно загружен. Теперь выберите маппинг полей.'))
            return redirect('data_import:preview_import', pk=import_file.pk)
    else:
        form = ImportFileForm(user=request.user)
    
    context = {
        'form': form,
        'mappings': PropertyImportMapping.objects.all()[:5],  # Показываем доступные маппинги
    }
    
    return render(request, 'data_import/upload_file.html', context)


@login_required
@user_passes_test(is_staff_user)
def preview_import(request, pk):
    """Предварительный просмотр данных импорта"""
    import_file = get_object_or_404(ImportFile, pk=pk)
    
    if request.method == 'POST':
        form = ImportPreviewForm(request.POST)
        if form.is_valid():
            mapping = form.cleaned_data['mapping']
            
            # Парсим файл с выбранным маппингом
            parser = ExcelParser(import_file.file.path, mapping)
            parse_result = parser.parse_file()
            
            if parse_result['success']:
                # Сохраняем маппинг и распарсенные данные
                import_file.mapping = mapping
                import_file.parsed_data = parse_result['data'][:10]  # Только первые 10 строк для превью
                import_file.total_rows = parse_result['total_rows']
                import_file.save()
                
                context = {
                    'import_file': import_file,
                    'form': form,
                    'preview_data': parse_result['data'][:10],
                    'headers': parse_result['headers'],
                    'total_rows': parse_result['total_rows'],
                    'mapping': mapping,
                    'mapping_display': mapping.field_mapping if mapping else {},
                }
                
                return render(request, 'data_import/preview_import.html', context)
            else:
                messages.error(request, _('Ошибка парсинга файла: {}').format('; '.join(parse_result['errors'])))
    else:
        form = ImportPreviewForm()
    
    context = {
        'import_file': import_file,
        'form': form,
    }
    
    return render(request, 'data_import/preview_import.html', context)


@login_required
@user_passes_test(is_staff_user)
@require_POST
def process_import(request, pk):
    """Запуск обработки импорта"""
    import_file = get_object_or_404(ImportFile, pk=pk)
    
    if import_file.status != 'uploaded':
        messages.warning(request, _('Файл уже обрабатывается или обработан'))
        return redirect('data_import:import_detail', pk=pk)
    
    # Запускаем обработку в отдельном потоке
    def run_import():
        processor = ImportProcessor(import_file)
        processor.process_import()
    
    thread = threading.Thread(target=run_import)
    thread.daemon = True
    thread.start()
    
    messages.success(request, _('Импорт запущен. Вы можете отслеживать прогресс на странице деталей.'))
    return redirect('data_import:import_detail', pk=pk)


@login_required
@user_passes_test(is_staff_user)
def import_list(request):
    """Список всех импортов"""
    imports = ImportFile.objects.select_related('uploaded_by').order_by('-created_at')
    
    # Фильтрация
    status_filter = request.GET.get('status')
    if status_filter:
        imports = imports.filter(status=status_filter)
    
    search = request.GET.get('search')
    if search:
        imports = imports.filter(
            Q(name__icontains=search) | 
            Q(uploaded_by__username__icontains=search)
        )
    
    # Пагинация
    paginator = Paginator(imports, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Обработка массовых действий
    if request.method == 'POST':
        bulk_form = BulkActionForm(request.POST)
        if bulk_form.is_valid():
            action = bulk_form.cleaned_data['action']
            file_ids = bulk_form.cleaned_data['selected_files']
            
            if action == 'delete':
                deleted_count = ImportFile.objects.filter(pk__in=file_ids).delete()[0]
                messages.success(request, _('Удалено {} файлов').format(deleted_count))
            elif action == 'process':
                for file_id in file_ids:
                    try:
                        import_file = ImportFile.objects.get(pk=file_id, status='uploaded')
                        # Запускаем в отдельном потоке
                        def run_import():
                            processor = ImportProcessor(import_file)
                            processor.process_import()
                        thread = threading.Thread(target=run_import)
                        thread.daemon = True
                        thread.start()
                    except ImportFile.DoesNotExist:
                        continue
                messages.success(request, _('Запущена обработка {} файлов').format(len(file_ids)))
                
            return redirect('data_import:import_list')
    else:
        bulk_form = BulkActionForm()
    
    context = {
        'page_obj': page_obj,
        'bulk_form': bulk_form,
        'status_filter': status_filter,
        'search': search,
    }
    
    return render(request, 'data_import/import_list.html', context)


@login_required
@user_passes_test(is_staff_user)
def import_detail(request, pk):
    """Детальная страница импорта"""
    import_file = get_object_or_404(ImportFile, pk=pk)
    logs = import_file.logs.order_by('-created_at')
    
    # Пагинация логов
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    logs_page = paginator.get_page(page_number)
    
    context = {
        'import_file': import_file,
        'logs_page': logs_page,
        'validation_errors': import_file.validation_errors or [],
    }
    
    return render(request, 'data_import/import_detail.html', context)


@login_required
@user_passes_test(is_staff_user)
def mapping_list(request):
    """Список маппингов полей"""
    mappings = PropertyImportMapping.objects.order_by('-is_default', 'name')
    
    context = {
        'mappings': mappings,
    }
    
    return render(request, 'data_import/mapping_list.html', context)


@login_required
@user_passes_test(is_staff_user)
def mapping_create(request):
    """Создание нового маппинга"""
    if request.method == 'POST':
        form = MappingForm(request.POST)
        if form.is_valid():
            mapping = form.save()
            messages.success(request, _('Маппинг "{}" успешно создан').format(mapping.name))
            return redirect('data_import:mapping_list')
    else:
        form = MappingForm()
    
    context = {
        'form': form,
        'available_fields': PropertyImportMapping.PROPERTY_FIELDS,
    }
    
    return render(request, 'data_import/mapping_form.html', context)


@login_required
@user_passes_test(is_staff_user)
def mapping_edit(request, pk):
    """Редактирование маппинга"""
    mapping = get_object_or_404(PropertyImportMapping, pk=pk)
    
    if request.method == 'POST':
        form = MappingForm(request.POST, instance=mapping)
        if form.is_valid():
            mapping = form.save()
            messages.success(request, _('Маппинг "{}" успешно обновлен').format(mapping.name))
            return redirect('data_import:mapping_list')
    else:
        form = MappingForm(instance=mapping)
    
    context = {
        'form': form,
        'mapping': mapping,
        'available_fields': PropertyImportMapping.PROPERTY_FIELDS,
    }
    
    return render(request, 'data_import/mapping_form.html', context)


@login_required
@user_passes_test(is_staff_user)
@require_POST
def mapping_delete(request, pk):
    """Удаление маппинга"""
    mapping = get_object_or_404(PropertyImportMapping, pk=pk)
    name = mapping.name
    mapping.delete()
    messages.success(request, _('Маппинг "{}" удален').format(name))
    return redirect('data_import:mapping_list')


@login_required
@user_passes_test(is_staff_user)
@require_http_methods(["GET"])
def import_status(request, pk):
    """API endpoint для получения статуса импорта (для AJAX)"""
    import_file = get_object_or_404(ImportFile, pk=pk)
    
    data = {
        'status': import_file.status,
        'total_rows': import_file.total_rows,
        'processed_rows': import_file.processed_rows,
        'successful_rows': import_file.successful_rows,
        'failed_rows': import_file.failed_rows,
        'success_rate': import_file.success_rate,
        'created_at': import_file.created_at.isoformat() if import_file.created_at else None,
        'started_at': import_file.started_at.isoformat() if import_file.started_at else None,
        'completed_at': import_file.completed_at.isoformat() if import_file.completed_at else None,
    }
    
    return JsonResponse(data)


@login_required
@user_passes_test(is_staff_user)
def download_template(request):
    """Скачивание шаблона Excel файла"""
    import openpyxl
    from django.http import HttpResponse
    from io import BytesIO
    
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Шаблон импорта недвижимости"
    
    # Заголовки (первая строка)
    headers = [
        'ID объекта', 'Название', 'Цена продажи USD', 'Цена продажи THB', 
        'Цена продажи RUB', 'Аренда USD', 'Аренда THB', 'Аренда RUB',
        'Общая площадь', 'Жилая площадь', 'Площадь участка', 'Спальни',
        'Ванные', 'Этаж', 'Всего этажей', 'Статус', 'Тип сделки',
        'Широта', 'Долгота', 'С мебелью', 'Бассейн', 'Парковка',
        'Охрана', 'Спортзал', 'Год постройки'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Примеры данных (вторая строка)
    example_data = [
        'VS001', 'Вилла у моря', 500000, 18000000, 45000000,
        2000, 72000, 180000, 300, 250, 500, 3, 2, 1, 2,
        'available', 'sale', 7.8, 98.3, 'true', 'true', 'true',
        'true', 'false', 2020
    ]
    
    for col, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    # Сохраняем в BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Создаем HTTP ответ
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="property_import_template.xlsx"'
    
    return response
